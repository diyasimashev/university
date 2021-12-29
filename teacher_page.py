import openpyxl
import pandas as pd


def teacher_greeting():
    print('\nWelcome, dear teacher!')


def teacher_menu():
    x = int(input("""\nIn order to get the information you are interested in, enter the number of the corresponding menu: \n1 - Exam Schedule\n2 - Maximum Grade/Minimum Grade\n3 - Schedule\n4 - Students\n5 - Students' Attendance\n6 - Students' Grades\nIf you want to terminate the program, enter 7: """))
    if x == 1:
        z = pd.read_excel('director.xlsx', sheet_name='sheet2')
        x1 = int(input("""\nTo get information about the exams, enter the subject number: \n1 - Algorithms and Data Structures\n2 - Human-Computer Interaction\n3 - Mathematics\n4 - Programming Languages\n5 - The English Language\n6 - The German Language: """))
        if x1 == 1:
            x1 = z[z["Subject"] == 'Algorithms and Data Structures']
            print('\n', x1)
            comeback()
        if x1 == 2:
            x1 = z[z["Subject"] == 'Human-Computer Interaction']
            print('\n', x1)
            comeback()
        if x1 == 3:
            x1 = z[z["Subject"] == 'Mathematics']
            print('\n', x1)
            comeback()
        if x1 == 4:
            x1 = z[z["Subject"] == 'Programming Languages']
            print('\n', x1)
            comeback()
        if x1 == 5:
            x1 = z[z["Subject"] == 'The English Language']
            print('\n', x1)
            comeback()
        if x1 == 6:
            x1 = z[z["Subject"] == 'The German Language']
            print('\n', x1)
            comeback()
        else:
            print("\nThe data you entered was not found. Try again.")
            comeback()
    elif x == 2:
        z = pd.read_excel('teacher.xlsx', sheet_name='sheet1', usecols=['Student', 'Average Grade'])
        x1 = int(input("""\nTo get information about a student with maximum/minimum grade, enter the subject number: \n1 - Mathematics\n2 - Mathematics(Analysis): """))
        if x1 == 1:
            x2 = int(input("""\nDo you want to see the name of student with: \n1 - Maximum Grade\n2 - Minimum Grade: """))
            if x2 == 1:
                x3 = z[z['Average Grade'] == z['Average Grade'].max()]
                print('\n', x3)
                comeback()
            elif x2 == 2:
                x3 = z[z['Average Grade'] == z['Average Grade'].min()]
                print('\n', x3)
                comeback()
            else:
                print("The data you entered was not found. Try again.")
                comeback()
        elif x1 == 2:
            print("\nThere are no grades yet for this subject")
            comeback()
        else:
            print("\nThe data you entered was not found. Try again.")
            comeback()
    elif x == 3:
        x1 = int(input("""\nTo get information about the schedule, enter the day number: \n1 - Monday\n2 - Tuesday\n3 - Wednesday\n4 - Thursday\n5 - Friday\n6 - Saturday: """))
        if x1 == 1:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Monday'])
            print('\n', z[z.Monday.notna()])
            comeback()
        elif x1 == 2:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Tuesday'])
            print('\n', z[z.Tuesday.notna()])
            comeback()
        elif x1 == 3:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Wednesday'])
            print('\n', z[z.Wednesday.notna()])
            comeback()
        elif x1 == 4:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Thursday'])
            print('\n', z[z.Thursday.notna()])
            comeback()
        elif x1 == 5:
            print("\nYou don't have any lessons on this day. Enjoy the day off!")
            comeback()
        elif x1 == 6:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Saturday'])
            print('\n', z[z.Saturday.notna()])
            comeback()
        else:
            print("\nThe data you entered was not found. Try again.")
            comeback()
    elif x == 4:
        x1 = int(input("""\nWhat information are you interested in?\n1 - Total Number of Students\n2 - Find a Student: """))
        if x1 == 1:
            z = openpyxl.open('teacher.xlsx', read_only=True)
            z.active = 0
            x2 = z.active
            for i in range(1, x2.max_row + 1):
                x3 = x2[i][0].value
                print('\n', x3)
            x4 = pd.read_excel('teacher.xlsx', sheet_name='sheet1')
            print("\nTotal number of students: " + str(len(x4)))
            comeback()
        elif x1 == 2:
            z = pd.read_excel('teacher.xlsx', sheet_name='sheet1')
            x2 = input("\nIn order to find a student, enter his(her) first and last name (observe the order!): ")
            x3 = z[z["Student"] == x2]
            print(x3)
            comeback()
        else:
            print("\nThe data you entered was not found. Try again.")
            comeback()
    elif x == 5:
        z = pd.read_excel('teacher.xlsx', sheet_name='sheet3')
        x1 = input("""\nTo view student's attendance, enter the student's last name and first name (observe the order): """)
        x2 = z[z["Student"] == x1]
        x3 = z.index[z["Student"] == x1]
        print(x2)
        x4 = int(input("\nChoose the next move: \n1 - Mark a Student\n2 - Comeback to Menu: "))
        if x4 == 1:
            x5 = input("\nEnter the date (for example, 5 September): ")
            x6 = int(input("\nMark a student: \n1 - attended the class\n2 - was absent from class: "))
            if x6 == 1:
                x6 = 'AT'
                x7 = openpyxl.load_workbook('teacher.xlsx')
                x7.active = 2
                x8 = x7.active
                for i in range(x8.max_column + 1, x8.max_column + 2):
                    if i is None:
                        continue
                    x8.cell(row=1, column=i).value = x5
                    x8.cell(row=x3[0] + 2, column=i).value = x6
                    x7.save('teacher.xlsx')
                comeback()
            elif x6 == 2:
                x6 = 'AB'
                x7 = openpyxl.load_workbook('teacher.xlsx')
                x7.active = 2
                x8 = x7.active
                for i in range(x8.max_column + 1, x8.max_column + 2):
                    if i is None:
                        continue
                    x8.cell(row=1, column=i).value = x5
                    x8.cell(row=x3[0] + 2, column=i).value = x6
                    x7.save('teacher.xlsx')
                comeback()
        elif x4 == 2:
            teacher_menu()
        else:
            print("The data you entered was not found. Try again.")
            comeback()
    elif x == 6:
        z = pd.read_excel('teacher.xlsx', sheet_name='sheet1')
        x1 = input("""\nTo view student's grades, enter the student's last name and first name (observe the order!): """)
        x2 = z[z["Student"] == x1]
        x3 = z.index[z["Student"] == x1]
        print(x2)
        x4 = int(input("\nChoose the next move: \n1 - Add a Grade\n2 - Change a Grade\n3 - Delete a grade\n4 - Comeback to Menu: "))
        if x4 == 1:
            x5 = int(input("\nEnter the grade: "))
            x6 = openpyxl.load_workbook('teacher.xlsx')
            x6.active = 0
            x7 = x6.active
            x7.cell(row=x3[0] + 2, column=4).value = int(x5)
            x6.save('teacher.xlsx')
            comeback()
        elif x4 == 2:
            x5 = int(input("\nThe grade for which date you want to change: \n1 - 1 September\n2 - 2 September\n3 - 3 September: "))
            if x5 == 1:
                x6 = int(input("\nEnter the grade: "))
                x7 = openpyxl.load_workbook('teacher.xlsx')
                x7.active = 0
                x8 = x7.active
                x8.cell(row=x3[0] + 2, column=2).value = x6
                x7.save('teacher.xlsx')
                comeback()
            elif x5 == 2:
                x6 = int(input("\nEnter the grade: "))
                x7 = openpyxl.load_workbook('teacher.xlsx')
                x7.active = 0
                x8 = x7.active
                x8.cell(row=x3[0] + 2, column=3).value = x6
                x7.save('teacher.xlsx')
                comeback()
            elif x5 == 3:
                x6 = int(input("\nEnter the grade: "))
                x7 = openpyxl.load_workbook('teacher.xlsx')
                x7.active = 0
                x8 = x7.active
                x8.cell(row=x3[0] + 2, column=4).value = x6
                x7.save('teacher.xlsx')
                comeback()
            else:
                print("\nThe data you entered was not found. Try again.")
                comeback()
        elif x4 == 3:
            x5 = int(input("\nThe grade for which date you want to delete: \n1 - 1 September\n2 - 2 September\n3 - 3 September: "))
            if x5 == 1:
                x6 = openpyxl.load_workbook('teacher.xlsx')
                x6.active = 0
                x7 = x6.active
                x7.cell(row=x3[0] + 2, column=2).value = None
                x6.save('teacher.xlsx')
                comeback()
            elif x5 == 2:
                x6 = openpyxl.load_workbook('teacher.xlsx')
                x6.active = 0
                x7 = x6.active
                x7.cell(row=x3[0] + 2, column=3).value = None
                x6.save('teacher.xlsx')
                comeback()
            elif x5 == 3:
                x6 = openpyxl.load_workbook('teacher.xlsx')
                x6.active = 0
                x7 = x6.active
                x7.cell(row=x3[0] + 2, column=4).value = None
                x6.save('teacher.xlsx')
                comeback()
            else:
                print("\nThe data you entered was not found. Try again.")
                comeback()
        elif x4 == 4:
            teacher_menu()
        else:
            print("\nThe data you entered was not found. Try again.")
            comeback()
    elif x == 7:
        for i in range(1):
            print("\nThe work of the program is completed. We are waiting for your next appearance.")
            break
    else:
        teacher_menu()


def comeback():
    q = input('To return to the menu, enter "c": ')
    if q == 'c':
        for k in range(1):
            teacher_menu()
            break
    else:
        for k in range(1):
            comeback()
