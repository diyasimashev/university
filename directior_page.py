import pandas as pd
import openpyxl


def director_greeting():
    print('\nWelcome, dear director!')


def director_menu():
    x = int(input("""\nIn order to get the information you are interested in, enter the number of the corresponding menu: \n1 - Add\n2 - Remove\n3 - Schedule\n4 - Students\n5 - Subjects\n6 - Teachers\nIf you want to terminate the program, enter 7:  """))
    if x == 1:
        x1 = int(input("""\nWho do you want to add? \n1 - Student\n2 - Teacher: """))
        if x1 == 1:
            adds()
        elif x1 == 2:
            addt()
        else:
            print("The data you entered was not found. Try again.")
            comeback()
    elif x == 2:
        x1 = int(input("""\nWho do you want to remove? \n1 - Student\n2 - Teacher: """))
        if x1 == 1:
            rs()
        elif x1 == 2:
            rt()
        else:
            print("\nThe data you entered was not found. Try again.")
            comeback()
    elif x == 3:
        b = int(input("""To get the schedule, enter the day number: \n1 - Monday\n2 - Tuesday\n3 - Wednesday\n4 - Thursday\n5 - Friday\n6 - Saturday: """))
        if b == 1:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Monday'])
            print(z[z.Monday.notna()])
            comeback()
        if b == 2:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Tuesday'])
            print(z[z.Tuesday.notna()])
            comeback()
        if b == 3:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Wednesday'])
            print(z[z.Wednesday.notna()])
            comeback()
        if b == 4:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Thursday'])
            print(z[z.Thursday.notna()])
            comeback()
        if b == 5:
            print("You don't have any lessons on this day.")
            comeback()
        if b == 6:
            z = pd.read_excel('director.xlsx', sheet_name='sheet3', usecols=['Start time of lessons', 'Saturday'])
            print(z[z.Saturday.notna()])
            comeback()
        else:
            print("The item number you entered is not in the database.")
            director_menu()
    elif x == 4:
        x1 = int(input("""\nWhat information are you interested in? \n1 - Name of the Student with the Maximum Grade\n2 - Name of the Student with the Minimum Grade\n3 - Student's Attendance\n4 - Student's Grades\n5 - Total Number of Students: """))
        if x1 == 1:
            z = pd.read_excel('teacher.xlsx', sheet_name='sheet1', usecols=['Student', 'Average Grade'])
            x2 = int(input("""\nTo get information about a student with maximum grade, enter the subject number: \n1 - Algorithms and Data Structure\n2 - Human-Computer Interaction\n3 - Mathematics\n4 - Programming Languages\n5 - The English Language\n6 - The German Language: """))
            if 1 <= x2 <= 2 or 4 <= x2 <= 6:
                print("\nThe teacher hasn't graded anyone yet")
                comeback()
            elif x2 == 3:
                x3 = z[z['Average Grade'] == z['Average Grade'].max()]
                print('\n', x3)
                comeback()
            else:
                print("\nThe data you entered was not found. Try again.")
                comeback()
        elif x1 == 2:
            z = pd.read_excel('teacher.xlsx', sheet_name='sheet1', usecols=['Student', 'Average Grade'])
            x2 = int(input("""\nTo get information about a student with minimum grade, enter the subject number: \n1 - Algorithms and Data Structure\n2 - Human-Computer Interaction\n3 - Mathematics\n4 - Programming Languages\n5 - The English Language\n6 - The German Language: """))
            if 1 <= x2 <= 2 or 4 <= x2 <= 6:
                print("\nThe teacher hasn't graded anyone yet")
                comeback()
            elif x2 == 3:
                x3 = z[z['Average Grade'] == z['Average Grade'].min()]
                print('\n', x3)
                comeback()
            else:
                print("\nThe data you entered was not found. Try again.")
                comeback()
        elif x1 == 3:
            x2 = int(input("""\nTo get information about student's attendance, enter the subject number: \n1 - Algorithms and Data Structure\n2 - Human-Computer Interaction\n3 - Mathematics\n4 - Programming Languages\n5 - The English Language\n6 - The German Language: """))
            if 1 <= x2 <= 2 or 4 <= x2 <= 6:
                print("\nThe teacher hasn't marked anyone yet.")
                comeback()
            elif x2 == 3:
                z = pd.read_excel('teacher.xlsx', sheet_name='sheet3')
                x1 = input("""\nTo view student's attendance, enter the student's last name and first name (observe the order): """)
                x2 = z[z["Student"] == x1]
                print(x2)
                comeback()
            else:
                print("\nThe data you entered was not found. Try again.")
                comeback()
        elif x1 == 4:
            x2 = int(input("""\nTo get information about student's grades, enter the subject number: \n1 - Algorithms and Data Structure\n2 - Human-Computer Interaction\n3 - Mathematics\n4 - Programming Languages\n5 - The English Language\n6 - The German Language: """))
            if 1 <= x2 <= 2 or 4 <= x2 <= 6:
                print("\nThe teacher hasn't graded anyone yet")
                comeback()
            elif x2 == 3:
                z = pd.read_excel('teacher.xlsx', sheet_name='sheet1')
                x3 = input("""\nTo view student's grades, enter the student's last name and first name (observe the order!): """)
                x4 = z[z["Student"] == x3]
                print(x4)
                comeback()
            else:
                print("\nThe data you entered was not found. Try again.")
                comeback()
        elif x1 == 5:
            z = openpyxl.open('teacher.xlsx', read_only=True)
            z.active = 0
            x2 = z.active
            for i in range(1, x2.max_row + 1):
                x3 = x2[i][0].value
                print('\n', x3)
            x4 = pd.read_excel('list_of_students.xlsx', sheet_name='Sheet1')
            print("\nTotal number of students: " + str(len(x4)))
            comeback()
        else:
            print("The data you entered was not found. Try again.")
            comeback()
    elif x == 5:
        z = openpyxl.open('director.xlsx', read_only=True)
        z.active = 0
        x1 = z.active
        for row in range(1, x1.max_row + 1):
            x2 = x1[row][0].value
            print('\n', x2)
        df = pd.read_excel('director.xlsx', sheet_name='sheet1')
        print("\nTotal number of subjects: " + str(len(df)))
        comeback()
    elif x == 6:
        x1 = int(input("\nWhat information are you interested in? \n1 - Total number of teachers\n2 - Searching a teacher: "))
        if x1 == 1:
            z = openpyxl.open('list_of_teachers.xlsx', read_only=True)
            z.active = 0
            c = z.active
            for i in range(1, c.max_row + 1):
                d = c[i][0].value
                print(d)
            df = pd.read_excel('list_of_teachers.xlsx', sheet_name='Sheet1')
            print("\nTotal number of teachers: " + str(len(df)))
            comeback()
        elif x1 == 2:
            z = pd.read_excel('list_of_teachers.xlsx', sheet_name='Sheet1')
            c = input("\nEnter a teacher last name and first name, please keep an order: ")
            d = z[z["Teacher"] == c]
            print(d)
            comeback()
        else:
            print("\nThe item number you entered is not in the database.")
            director_menu()
    elif x == 7:
        for i in range(1):
            print("\nThe work of the program is completed. We are waiting for your next appearance.")
            break
    else:
        director_menu()


def addt():
    x1 = input("\nEnter the teacher's last name and first name ( follow the sequence ): ")
    x2 = input("Enter the name of the lesson that the teacher is teaching:  ")
    z = openpyxl.load_workbook('list_of_teachers.xlsx')
    z.active = 0
    x3 = z.active
    if x1 is None:
        print("\nYou must enter the teacher's last name and first name, otherwise it is impossible to save the data.")
        addt()
    elif x2 is None:
        print("\nYou must Enter the name of the lesson that the teacher is teaching, otherwise it is impossible to save the data.")
        addt()
    else:
        for j in range(x3.max_row + 1, x3.max_row + 2):
            if j is None:
                continue
            x3.cell(row=j, column=1).value = x1
            x3.cell(row=j, column=2).value = x2
            z.save('list_of_teachers.xlsx')
        print("\nThe data has been saved. Registration was successful.")
        comeback()


def adds():
    x1 = input("\nEnter the student's last name and first name ( follow the sequence ): ")
    z = openpyxl.load_workbook('list_of_students.xlsx')
    z.active = 0
    x2 = z.active
    if x1 is None:
        print("\nYou must enter the student's last name and first name, otherwise it is impossible to save the data.")
        adds()
    else:
        for j in range(x2.max_row + 1, x2.max_row + 2):
            if j is True:
                continue
            x2.cell(row=j, column=1).value = x1
            z.save('list_of_students.xlsx')
        print("\nThe data has been saved. Registration was successful.")
        comeback()


def rs():
    x1 = input("Enter the student's last name and first name ( follow the sequence ): ")
    z = pd.read_excel('list_of_students.xlsx', sheet_name='Sheet1')
    z = z[z['Student'].apply(lambda x: x != x1)]
    z.to_excel('list_of_students.xlsx', index=False)
    comeback()


def rt():
    x1 = input("\nEnter the teacher's last name and first name ( follow the sequence ): ")
    z = pd.read_excel('list_of_teachers.xlsx', sheet_name='Sheet1')
    z = z[z['Teacher'].apply(lambda x: x != x1)]
    z.to_excel('list_of_teachers.xlsx', index=False)
    comeback()


def comeback():
    x = input('To return to the menu, enter "c": ')
    if x == 'c':
        for i in range(1):
            director_menu()
            break
    else:
        for i in range(1):
            comeback()
